"""
La Sastrería - Sistema de Gestión
Fundador: Erick Gutierrez | Administradora: Keila Gutierrez
"""
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from io import BytesIO
import hashlib, os, json, psycopg2, psycopg2.extras, openpyxl, smtplib, functools
from datetime import datetime, date, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sastreria_erick_keila_2024_xK9mP2qL')
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres.rmczinwjgxkfgffdsaiy:Ca83fd59f1.@aws-0-us-west-2.pooler.supabase.com:5432/postgres')
EXPORTS_PATH = '/tmp/exports'

USERS = {
    'erick': {'password': 'c98583b2cb9f8dcd129c0cda6913a8ed1db835a877842ee208569b50b97f46ee', 'name': 'Erick Gutierrez', 'role': 'owner', 'display': 'Dueño & Fundador'},
    'keila': {'password': 'd6730f9706f88db370fef218422d05d16e3f94b9bbcf5dfa239cf21ff4ce8a11', 'name': 'Keila Gutierrez', 'role': 'admin', 'display': 'Administradora'}
}

def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor, connect_timeout=10)
    return conn

def init_db():
    os.makedirs(EXPORTS_PATH, exist_ok=True)
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS clientes (id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, telefono TEXT, correo TEXT, nit TEXT, direccion TEXT, notas TEXT, fecha_registro TEXT NOT NULL, activo INTEGER DEFAULT 1)''')
    c.execute('''CREATE TABLE IF NOT EXISTS ordenes (id SERIAL PRIMARY KEY, numero_orden TEXT UNIQUE NOT NULL, cliente_id INTEGER NOT NULL, fecha_orden TEXT NOT NULL, fecha_entrega TEXT NOT NULL, estado TEXT DEFAULT 'pendiente', total REAL DEFAULT 0, notas_adicionales TEXT, usuario_registro TEXT, fecha_creacion TEXT NOT NULL, correo_enviado INTEGER DEFAULT 0, FOREIGN KEY (cliente_id) REFERENCES clientes(id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS prendas_orden (id SERIAL PRIMARY KEY, orden_id INTEGER NOT NULL, tipo_prenda TEXT NOT NULL, descripcion_servicio TEXT NOT NULL, cantidad INTEGER DEFAULT 1, precio_tipo TEXT DEFAULT 'fijo', precio_unitario REAL NOT NULL, subtotal REAL NOT NULL, entregada INTEGER DEFAULT 0, fecha_entrega_real TEXT, FOREIGN KEY (orden_id) REFERENCES ordenes(id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS catalogo_servicios (id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, descripcion TEXT, precio_base REAL, precio_tipo TEXT DEFAULT 'fijo', activo INTEGER DEFAULT 1)''')
    c.execute('''CREATE TABLE IF NOT EXISTS configuracion (clave TEXT PRIMARY KEY, valor TEXT)''')
    for clave, valor in [('empresa_nombre','La Sastreria'),('smtp_servidor','smtp.gmail.com'),('smtp_puerto','587'),('smtp_usuario',''),('smtp_password',''),('moneda','Q')]:
        c.execute('INSERT INTO configuracion VALUES (%s,%s) ON CONFLICT (clave) DO NOTHING', (clave, valor))
    for nombre, desc, precio, tipo in [('Pantalon - Ruedo','Subir o bajar ruedo',25.0,'fijo'),('Pantalon - Entrada','Ajuste cintura',35.0,'fijo'),('Camisa - Manga','Acortar mangas',30.0,'fijo'),('Vestido - Arreglo','Arreglo general',0.0,'variable'),('Traje - Confeccion','Confeccion a medida',0.0,'variable'),('Falda - Ruedo','Subir o bajar ruedo',25.0,'fijo'),('Zipper - Cambio','Cambio de zipper',30.0,'fijo'),('Botones - Costura','Costura de botones',10.0,'fijo')]:
        c.execute('INSERT INTO catalogo_servicios (nombre,descripcion,precio_base,precio_tipo) SELECT %s,%s,%s,%s WHERE NOT EXISTS (SELECT 1 FROM catalogo_servicios WHERE nombre=%s)', (nombre,desc,precio,tipo,nombre))
    conn.commit()
    conn.close()
    print("DB lista")

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def generate_order_number():
    conn = get_db()
    c = conn.cursor()
    prefix = f"ORD-{date.today().strftime('%Y%m')}-"
    c.execute("SELECT numero_orden FROM ordenes WHERE numero_orden LIKE %s ORDER BY id DESC LIMIT 1", (prefix+'%',))
    last = c.fetchone()
    conn.close()
    if last:
        return f"{prefix}{int(last['numero_orden'].split('-')[-1])+1:04d}"
    return f"{prefix}0001"

def get_config():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT clave, valor FROM configuracion")
    config = {r['clave']:r['valor'] for r in c.fetchall()}
    conn.close()
    return config

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'usuario' in session else url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        usuario = request.form.get('usuario','').strip().lower()
        password_hash = hashlib.sha256(request.form.get('password','').encode()).hexdigest()
        if usuario in USERS and USERS[usuario]['password'] == password_hash:
            session['usuario'] = usuario
            session['nombre'] = USERS[usuario]['name']
            session['rol'] = USERS[usuario]['role']
            session['display'] = USERS[usuario]['display']
            return redirect(url_for('dashboard'))
        error = 'Usuario o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as total FROM clientes WHERE activo=1")
    total_clientes = c.fetchone()['total']
    c.execute("SELECT COUNT(*) as total FROM ordenes WHERE estado='pendiente'")
    ordenes_pendientes = c.fetchone()['total']
    c.execute("SELECT COUNT(*) as total FROM ordenes WHERE estado='entregado'")
    ordenes_entregadas = c.fetchone()['total']
    c.execute("SELECT COALESCE(SUM(total),0) as total FROM ordenes WHERE estado='entregado'")
    ingresos_total = c.fetchone()['total']
    hoy = date.today().isoformat()
    en_7_dias = (date.today() + timedelta(days=7)).isoformat()
    c.execute("SELECT o.*, cl.nombre as cliente_nombre FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id WHERE o.fecha_entrega BETWEEN %s AND %s AND o.estado='pendiente' ORDER BY o.fecha_entrega LIMIT 10", (hoy, en_7_dias))
    proximas_entregas = c.fetchall()
    c.execute("SELECT o.*, cl.nombre as cliente_nombre FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id ORDER BY o.fecha_creacion DESC LIMIT 8")
    ordenes_recientes = c.fetchall()
    c.execute("SELECT COUNT(*) as total FROM ordenes WHERE fecha_entrega=%s AND estado='pendiente'", (hoy,))
    entregas_hoy = c.fetchone()['total']
    conn.close()
    return render_template('dashboard.html', total_clientes=total_clientes, ordenes_pendientes=ordenes_pendientes, ordenes_entregadas=ordenes_entregadas, ingresos_total=ingresos_total, proximas_entregas=proximas_entregas, ordenes_recientes=ordenes_recientes, entregas_hoy=entregas_hoy, hoy=hoy)

@app.route('/clientes')
@login_required
def clientes():
    busqueda = request.args.get('q','')
    conn = get_db()
    c = conn.cursor()
    if busqueda:
        c.execute("SELECT cl.*, COUNT(o.id) as total_ordenes, MAX(o.fecha_orden) as ultima_orden FROM clientes cl LEFT JOIN ordenes o ON cl.id=o.cliente_id WHERE cl.activo=1 AND (cl.nombre ILIKE %s OR cl.telefono ILIKE %s OR cl.nit ILIKE %s OR cl.correo ILIKE %s) GROUP BY cl.id ORDER BY cl.nombre", (f'%{busqueda}%',)*4)
    else:
        c.execute("SELECT cl.*, COUNT(o.id) as total_ordenes, MAX(o.fecha_orden) as ultima_orden FROM clientes cl LEFT JOIN ordenes o ON cl.id=o.cliente_id WHERE cl.activo=1 GROUP BY cl.id ORDER BY cl.nombre")
    lista = c.fetchall()
    conn.close()
    return render_template('clientes.html', clientes=lista, busqueda=busqueda)

@app.route('/clientes/nuevo', methods=['GET','POST'])
@login_required
def nuevo_cliente():
    if request.method == 'POST':
        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO clientes (nombre,telefono,correo,nit,direccion,notas,fecha_registro) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (request.form['nombre'], request.form.get('telefono',''), request.form.get('correo',''), request.form.get('nit',''), request.form.get('direccion',''), request.form.get('notas',''), date.today().isoformat()))
        cliente_id = c.fetchone()['id']
        conn.commit()
        conn.close()
        flash('Cliente registrado exitosamente', 'success')
        return redirect(url_for('detalle_cliente', id=cliente_id))
    return render_template('nuevo_cliente.html')

@app.route('/clientes/<int:id>')
@login_required
def detalle_cliente(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM clientes WHERE id=%s", (id,))
    cliente = c.fetchone()
    if not cliente:
        return redirect(url_for('clientes'))
    c.execute("SELECT o.*, COUNT(po.id) as total_prendas FROM ordenes o LEFT JOIN prendas_orden po ON o.id=po.orden_id WHERE o.cliente_id=%s GROUP BY o.id ORDER BY o.fecha_orden DESC", (id,))
    ordenes = c.fetchall()
    conn.close()
    return render_template('detalle_cliente.html', cliente=cliente, ordenes=ordenes)

@app.route('/clientes/<int:id>/editar', methods=['GET','POST'])
@login_required
def editar_cliente(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM clientes WHERE id=%s", (id,))
    cliente = c.fetchone()
    if request.method == 'POST':
        c.execute("UPDATE clientes SET nombre=%s,telefono=%s,correo=%s,nit=%s,direccion=%s,notas=%s WHERE id=%s",
            (request.form['nombre'], request.form.get('telefono',''), request.form.get('correo',''), request.form.get('nit',''), request.form.get('direccion',''), request.form.get('notas',''), id))
        conn.commit()
        conn.close()
        flash('Cliente actualizado', 'success')
        return redirect(url_for('detalle_cliente', id=id))
    conn.close()
    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/api/clientes/buscar')
@login_required
def api_buscar_clientes():
    q = request.args.get('q','')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id,nombre,telefono,correo,nit FROM clientes WHERE activo=1 AND nombre ILIKE %s LIMIT 10", (f'%{q}%',))
    resultados = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(resultados)

@app.route('/ordenes')
@login_required
def ordenes():
    estado = request.args.get('estado','todos')
    busqueda = request.args.get('q','')
    conn = get_db()
    c = conn.cursor()
    query = "SELECT o.*, cl.nombre as cliente_nombre, cl.telefono as cliente_tel, o.usuario_registro, COUNT(po.id) as total_prendas FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id LEFT JOIN prendas_orden po ON o.id=po.orden_id WHERE 1=1"
    params = []
    if estado != 'todos':
        query += " AND o.estado=%s"; params.append(estado)
    if busqueda:
        query += " AND (cl.nombre ILIKE %s OR o.numero_orden ILIKE %s)"; params.extend([f'%{busqueda}%',f'%{busqueda}%'])
    query += " GROUP BY o.id,cl.nombre,cl.telefono ORDER BY o.fecha_orden DESC"
    c.execute(query, params)
    lista = c.fetchall()
    conn.close()
    return render_template('ordenes.html', ordenes=lista, estado_filtro=estado, busqueda=busqueda)

@app.route('/ordenes/nueva', methods=['GET','POST'])
@login_required
def nueva_orden():
    if request.method == 'POST':
        data = request.get_json()
        conn = get_db()
        c = conn.cursor()
        numero_orden = data.get('numero_orden') or generate_order_number()
        cliente_id = data.get('cliente_id')
        if not cliente_id:
            c.execute("INSERT INTO clientes (nombre,telefono,correo,nit,fecha_registro) VALUES (%s,%s,%s,%s,%s) RETURNING id",
                (data['cliente_nombre'], data.get('cliente_tel',''), data.get('cliente_correo',''), data.get('cliente_nit',''), date.today().isoformat()))
            cliente_id = c.fetchone()['id']
        c.execute("INSERT INTO ordenes (numero_orden,cliente_id,fecha_orden,fecha_entrega,estado,total,notas_adicionales,usuario_registro,fecha_creacion) VALUES (%s,%s,%s,%s,'pendiente',%s,%s,%s,%s) RETURNING id",
            (numero_orden, cliente_id, data['fecha_orden'], data['fecha_entrega'], data['total'], data.get('notas',''), session['usuario'], datetime.now().isoformat()))
        orden_id = c.fetchone()['id']
        for p in data['prendas']:
            c.execute("INSERT INTO prendas_orden (orden_id,tipo_prenda,descripcion_servicio,cantidad,precio_tipo,precio_unitario,subtotal) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (orden_id, p['tipo'], p['descripcion'], p['cantidad'], p['precio_tipo'], p['precio_unitario'], p['subtotal']))
        conn.commit()
        c.execute("SELECT * FROM ordenes WHERE id=%s", (orden_id,))
        orden = dict(c.fetchone())
        c.execute("SELECT * FROM clientes WHERE id=%s", (cliente_id,))
        cliente = dict(c.fetchone())
        c.execute("SELECT * FROM prendas_orden WHERE orden_id=%s", (orden_id,))
        prendas = [dict(p) for p in c.fetchall()]
        conn.close()
        exportar_orden_excel(orden, cliente, prendas)
        if cliente.get('correo'):
            enviar_correo_orden(orden, cliente, prendas)
        return jsonify({'success': True, 'numero_orden': numero_orden, 'orden_id': orden_id})
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM catalogo_servicios WHERE activo=1 ORDER BY nombre")
    catalogo = [dict(s) for s in c.fetchall()]
    conn.close()
    return render_template('nueva_orden.html', catalogo=catalogo, numero_previo=generate_order_number(), fecha_hoy=date.today().isoformat())

@app.route('/ordenes/<int:id>')
@login_required
def detalle_orden(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT o.*, cl.nombre as cliente_nombre, cl.telefono as cliente_tel, cl.correo as cliente_correo, cl.nit as cliente_nit, cl.id as cliente_id FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id WHERE o.id=%s", (id,))
    orden = c.fetchone()
    if not orden:
        return redirect(url_for('ordenes'))
    c.execute("SELECT * FROM prendas_orden WHERE orden_id=%s ORDER BY id", (id,))
    prendas = c.fetchall()
    conn.close()
    return render_template('detalle_orden.html', orden=orden, prendas=prendas)

@app.route('/ordenes/<int:id>/editar', methods=['GET','POST'])
@login_required
def editar_orden(id):
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        data = request.get_json()
        c.execute("UPDATE ordenes SET fecha_entrega=%s,notas_adicionales=%s,total=%s WHERE id=%s", (data['fecha_entrega'], data.get('notas',''), data['total'], id))
        for p in data.get('prendas_actualizar',[]):
            c.execute("UPDATE prendas_orden SET tipo_prenda=%s,descripcion_servicio=%s,cantidad=%s,precio_unitario=%s,subtotal=%s WHERE id=%s", (p['tipo'],p['descripcion'],p['cantidad'],p['precio_unitario'],p['subtotal'],p['id']))
        for p in data.get('prendas_nuevas',[]):
            c.execute("INSERT INTO prendas_orden (orden_id,tipo_prenda,descripcion_servicio,cantidad,precio_tipo,precio_unitario,subtotal) VALUES (%s,%s,%s,%s,%s,%s,%s)", (id,p['tipo'],p['descripcion'],p['cantidad'],p['precio_tipo'],p['precio_unitario'],p['subtotal']))
        for pid in data.get('prendas_eliminar',[]):
            c.execute("DELETE FROM prendas_orden WHERE id=%s", (pid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    c.execute("SELECT o.*, cl.nombre as cliente_nombre FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id WHERE o.id=%s", (id,))
    orden = c.fetchone()
    c.execute("SELECT * FROM prendas_orden WHERE orden_id=%s", (id,))
    prendas = c.fetchall()
    c.execute("SELECT * FROM catalogo_servicios WHERE activo=1")
    catalogo = [dict(s) for s in c.fetchall()]
    conn.close()
    return render_template('editar_orden.html', orden=orden, prendas=prendas, catalogo=catalogo)

@app.route('/api/ordenes/<int:id>/estado', methods=['POST'])
@login_required
def cambiar_estado_orden(id):
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE ordenes SET estado=%s WHERE id=%s", (data.get('estado'), id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/prendas/<int:id>/entregar', methods=['POST'])
@login_required
def entregar_prenda(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE prendas_orden SET entregada=1, fecha_entrega_real=%s WHERE id=%s", (date.today().isoformat(), id))
    c.execute("SELECT orden_id FROM prendas_orden WHERE id=%s", (id,))
    row = c.fetchone()
    if row:
        oid = row['orden_id']
        c.execute("SELECT COUNT(*) as total, SUM(entregada) as entregadas FROM prendas_orden WHERE orden_id=%s", (oid,))
        counts = c.fetchone()
        if counts['total'] == counts['entregadas']:
            c.execute("UPDATE ordenes SET estado='entregado' WHERE id=%s", (oid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/calendario')
@login_required
def calendario():
    return render_template('calendario.html')

@app.route('/api/calendario/eventos')
@login_required
def api_eventos_calendario():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT o.id, o.numero_orden, o.fecha_entrega, o.estado, o.total, cl.nombre as cliente_nombre FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id ORDER BY o.fecha_entrega")
    ordenes = c.fetchall()
    conn.close()
    eventos = []
    for o in ordenes:
        color = '#f59e0b' if o['estado']=='pendiente' else '#10b981' if o['estado']=='entregado' else '#6b7280'
        eventos.append({'id':o['id'],'title':f"{o['numero_orden']} - {o['cliente_nombre']}","start":o['fecha_entrega'],'color':color,'estado':o['estado'],'total':float(o['total'])})
    return jsonify(eventos)

@app.route('/catalogo')
@login_required
def catalogo():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM catalogo_servicios ORDER BY nombre")
    servicios = [dict(s) for s in c.fetchall()]
    conn.close()
    return render_template('catalogo.html', servicios=servicios)

@app.route('/catalogo/nuevo', methods=['POST'])
@login_required
def nuevo_servicio():
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO catalogo_servicios (nombre,descripcion,precio_base,precio_tipo) VALUES (%s,%s,%s,%s)", (data['nombre'],data.get('descripcion',''),float(data.get('precio_base',0)),data.get('precio_tipo','fijo')))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/catalogo/<int:id>/editar', methods=['POST'])
@login_required
def editar_servicio(id):
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE catalogo_servicios SET nombre=%s,descripcion=%s,precio_base=%s,precio_tipo=%s,activo=%s WHERE id=%s", (data['nombre'],data.get('descripcion',''),float(data.get('precio_base',0)),data.get('precio_tipo','fijo'),data.get('activo',1),id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/ordenes/<int:id>/descargar')
@login_required
def descargar_orden(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT o.*, cl.nombre as cliente_nombre, cl.telefono as cliente_tel, cl.correo as cliente_correo, cl.nit as cliente_nit FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id WHERE o.id=%s", (id,))
    orden = dict(c.fetchone())
    c.execute("SELECT * FROM clientes WHERE id=%s", (orden['cliente_id'],))
    cliente = dict(c.fetchone())
    c.execute("SELECT * FROM prendas_orden WHERE orden_id=%s", (id,))
    prendas = [dict(p) for p in c.fetchall()]
    conn.close()
    filepath = exportar_orden_excel(orden, cliente, prendas)
    return send_file(filepath, as_attachment=True, download_name=f"Orden_{orden['numero_orden']}.xlsx")

@app.route('/exportar/clientes')
@login_required
def exportar_clientes():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT cl.*, COUNT(o.id) as total_ordenes, COALESCE(SUM(o.total),0) as total_gastado FROM clientes cl LEFT JOIN ordenes o ON cl.id=o.cliente_id WHERE cl.activo=1 GROUP BY cl.id ORDER BY cl.nombre")
    clientes = c.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    headers = ['ID','Nombre','Telefono','Correo','NIT','Direccion','Notas','Fecha Registro','Total Ordenes','Total Gastado Q']
    hf = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = Font(color="FFFFFF", bold=True)
    for rn, cl in enumerate(clientes, 2):
        for col, key in enumerate(['id','nombre','telefono','correo','nit','direccion','notas','fecha_registro','total_ordenes','total_gastado'], 1):
            val = cl[key]
            ws.cell(row=rn, column=col, value=round(float(val),2) if key in ['total_gastado'] else val)
    os.makedirs(EXPORTS_PATH, exist_ok=True)
    filepath = os.path.join(EXPORTS_PATH, f"Clientes_{date.today().isoformat()}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=f"Clientes_{date.today().isoformat()}.xlsx")

@app.route('/configuracion', methods=['GET','POST'])
@login_required
def configuracion():
    if session.get('rol') != 'owner':
        flash('Solo el dueno puede acceder', 'error')
        return redirect(url_for('dashboard'))
    config = get_config()
    if request.method == 'POST':
        conn = get_db()
        c = conn.cursor()
        for clave in ['smtp_servidor','smtp_puerto','smtp_usuario','smtp_password']:
            c.execute("UPDATE configuracion SET valor=%s WHERE clave=%s", (request.form.get(clave,''), clave))
        conn.commit()
        conn.close()
        flash('Configuracion guardada', 'success')
        config = get_config()
    return render_template('configuracion.html', config=config)

def exportar_orden_excel(orden, cliente, prendas):
    os.makedirs(EXPORTS_PATH, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orden"
    hf = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    af = PatternFill(start_color="c8a96e", end_color="c8a96e", fill_type="solid")
    sf = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    bd = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    ws.merge_cells('A1:G1'); ws['A1'] = 'LA SASTRERIA'; ws['A1'].font = Font(size=18,bold=True,color='1a1a2e'); ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2'); ws['A2'] = 'Erick Gutierrez - Fundador'; ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:G3'); ws['A3'] = f"ORDEN: {orden['numero_orden']}"; ws['A3'].font = Font(size=13,bold=True,color='c8a96e'); ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A5:C5'); ws['A5'] = 'CLIENTE'; ws['A5'].fill = hf; ws['A5'].font = sf
    ws['A6']='Nombre:'; ws['B6']=cliente.get('nombre','')
    ws['A7']='Telefono:'; ws['B7']=cliente.get('telefono','')
    ws['A8']='Correo:'; ws['B8']=cliente.get('correo','')
    ws['A9']='NIT:'; ws['B9']=cliente.get('nit','')
    ws.merge_cells('E5:G5'); ws['E5']='ORDEN'; ws['E5'].fill=hf; ws['E5'].font=sf
    ws['E6']='Fecha Orden:'; ws['F6']=orden.get('fecha_orden','')
    ws['E7']='Fecha Entrega:'; ws['F7']=orden.get('fecha_entrega','')
    ws['E8']='Estado:'; ws['F8']=str(orden.get('estado','')).upper()
    ws['E9']='Registrado por:'; ws['F9']=orden.get('usuario_registro','')
    for col, h in enumerate(['Tipo Prenda','Descripcion','Cant.','Tipo Precio','Precio Unit. Q','Subtotal Q','Entregada'],1):
        cell = ws.cell(row=11,column=col,value=h); cell.fill=hf; cell.font=sf; cell.alignment=Alignment(horizontal='center'); cell.border=bd
    for i, p in enumerate(prendas,12):
        ws.cell(row=i,column=1,value=p.get('tipo_prenda','')).border=bd
        ws.cell(row=i,column=2,value=p.get('descripcion_servicio','')).border=bd
        ws.cell(row=i,column=3,value=p.get('cantidad',1)).border=bd
        ws.cell(row=i,column=4,value=str(p.get('precio_tipo','fijo')).upper()).border=bd
        ws.cell(row=i,column=5,value=round(float(p.get('precio_unitario',0)),2)).border=bd
        ws.cell(row=i,column=6,value=round(float(p.get('subtotal',0)),2)).border=bd
        ws.cell(row=i,column=7,value='Si' if p.get('entregada') else 'No').border=bd
    lr = 12+len(prendas)
    ws.merge_cells(f'A{lr}:E{lr}'); ws[f'A{lr}']='TOTAL'; ws[f'A{lr}'].font=Font(bold=True,size=12); ws[f'A{lr}'].fill=af; ws[f'A{lr}'].alignment=Alignment(horizontal='right')
    ws[f'F{lr}']=round(float(orden.get('total',0)),2); ws[f'F{lr}'].font=Font(bold=True,size=12); ws[f'F{lr}'].fill=af
    if orden.get('notas_adicionales'):
        ws[f'A{lr+2}']='Notas:'; ws[f'B{lr+2}']=orden['notas_adicionales']
    for i,w in enumerate([25,35,12,14,18,14,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    filepath = os.path.join(EXPORTS_PATH, f"Orden_{orden['numero_orden']}.xlsx")
    wb.save(filepath)
    return filepath

def enviar_correo_orden(orden, cliente, prendas):
    try:
        config = get_config()
        smtp_user = config.get('smtp_usuario','')
        smtp_pass = config.get('smtp_password','')
        if not smtp_user or not smtp_pass:
            return False
        msg = MIMEMultipart()
        msg['From']=smtp_user; msg['To']=cliente['correo']; msg['Subject']=f"La Sastreria - Orden {orden['numero_orden']}"
        prendas_html=''.join([f"<tr><td>{p['tipo_prenda']}</td><td>{p['descripcion_servicio']}</td><td>{p['cantidad']}</td><td>Q{float(p['precio_unitario']):.2f}</td><td>Q{float(p['subtotal']):.2f}</td></tr>" for p in prendas])
        body=f"<html><body><h1>La Sastreria</h1><h2>Orden: {orden['numero_orden']}</h2><p>Cliente: {cliente['nombre']}</p><table border='1'>{prendas_html}</table><p>Entrega: {orden['fecha_entrega']}</p><p><b>TOTAL: Q{float(orden['total']):.2f}</b></p></body></html>"
        msg.attach(MIMEText(body,'html'))
        server=smtplib.SMTP(config.get('smtp_servidor','smtp.gmail.com'),int(config.get('smtp_puerto',587)))
        server.starttls(); server.login(smtp_user,smtp_pass); server.sendmail(smtp_user,cliente['correo'],msg.as_string()); server.quit()
        return True
    except Exception as e:
        print(f"Error correo: {e}")
        return False

@app.route('/ordenes/<int:id>/reenviar-correo', methods=['POST'])
@login_required
def reenviar_correo(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT o.*, cl.nombre as cliente_nombre FROM ordenes o JOIN clientes cl ON o.cliente_id=cl.id WHERE o.id=%s", (id,))
    orden = dict(c.fetchone())
    c.execute("SELECT * FROM clientes WHERE id=%s", (orden['cliente_id'],))
    cliente = dict(c.fetchone())
    c.execute("SELECT * FROM prendas_orden WHERE orden_id=%s", (id,))
    prendas = [dict(p) for p in c.fetchall()]
    conn.close()
    if not cliente.get('correo'):
        return jsonify({'success': False, 'error': 'Cliente sin correo'})
    return jsonify({'success': enviar_correo_orden(orden, cliente, prendas)})

try:
    init_db()
except Exception as e:
    import traceback
    print(f"Error DB: {e}")
    traceback.print_exc()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=os.environ.get('RENDER') is None, host='0.0.0.0', port=port)
